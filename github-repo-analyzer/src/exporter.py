"""
导出模块

支持多种格式导出分析结果：PDF、Excel、CSV、Markdown等
"""

import os
import json
import csv
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
from abc import ABC, abstractmethod

from rich.console import Console

console = Console()


class ExportStrategy(ABC):
    """导出策略基类"""
    
    @abstractmethod
    def export(self, data: Dict, filename: str) -> str:
        """导出数据"""
        pass
    
    @abstractmethod
    def get_extension(self) -> str:
        """获取文件扩展名"""
        pass


class CSVExporter(ExportStrategy):
    """CSV导出器"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def get_extension(self) -> str:
        return ".csv"
    
    def export(self, data: Dict, filename: str) -> str:
        """导出为CSV格式"""
        filepath = self.output_dir / f"{filename}{self.get_extension()}"
        
        # 导出多个CSV文件（每个数据集一个）
        exported_files = []
        
        # 导出Commit数据
        if 'commit_analysis' in data:
            commit_file = self._export_commits(data['commit_analysis'], filename)
            exported_files.append(commit_file)
        
        # 导出贡献者数据
        if 'contributor_analysis' in data:
            contrib_file = self._export_contributors(data['contributor_analysis'], filename)
            exported_files.append(contrib_file)
        
        # 导出Issue数据
        if 'issue_analysis' in data:
            issue_file = self._export_issues(data['issue_analysis'], filename)
            exported_files.append(issue_file)
        
        # 导出PR数据
        if 'pr_analysis' in data:
            pr_file = self._export_prs(data['pr_analysis'], filename)
            exported_files.append(pr_file)
        
        return ', '.join(exported_files)
    
    def _export_commits(self, commit_data: Dict, prefix: str) -> str:
        """导出Commit数据"""
        filepath = self.output_dir / f"{prefix}_commits.csv"
        
        # 提取月度数据
        monthly = commit_data.get('monthly_distribution', {}).get('distribution', {})
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['月份', 'Commit数量'])
            for month, count in monthly.items():
                writer.writerow([month, count])
        
        return str(filepath)
    
    def _export_contributors(self, contrib_data: Dict, prefix: str) -> str:
        """导出贡献者数据"""
        filepath = self.output_dir / f"{prefix}_contributors.csv"
        
        contributors = contrib_data.get('contributors', [])
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['用户名', '贡献数', '占比(%)'])
            for contrib in contributors:
                writer.writerow([
                    contrib.get('login', ''),
                    contrib.get('contributions', 0),
                    contrib.get('percentage', 0)
                ])
        
        return str(filepath)
    
    def _export_issues(self, issue_data: Dict, prefix: str) -> str:
        """导出Issue数据"""
        filepath = self.output_dir / f"{prefix}_issues.csv"
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['指标', '数值'])
            writer.writerow(['总Issues', issue_data.get('total_issues', 0)])
            writer.writerow(['开放Issues', issue_data.get('open_issues', 0)])
            writer.writerow(['关闭Issues', issue_data.get('closed_issues', 0)])
            writer.writerow(['平均关闭时间(天)', issue_data.get('avg_close_time_days', 0)])
        
        return str(filepath)
    
    def _export_prs(self, pr_data: Dict, prefix: str) -> str:
        """导出PR数据"""
        filepath = self.output_dir / f"{prefix}_prs.csv"
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['指标', '数值'])
            writer.writerow(['总PRs', pr_data.get('total_prs', 0)])
            writer.writerow(['开放PRs', pr_data.get('open_prs', 0)])
            writer.writerow(['合并PRs', pr_data.get('merged_prs', 0)])
            writer.writerow(['关闭PRs', pr_data.get('closed_prs', 0)])
            writer.writerow(['合并率(%)', pr_data.get('merge_rate', 0)])
        
        return str(filepath)


class ExcelExporter(ExportStrategy):
    """Excel导出器"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def get_extension(self) -> str:
        return ".xlsx"
    
    def export(self, data: Dict, filename: str) -> str:
        """导出为Excel格式"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import LineChart, BarChart, PieChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            console.print("[yellow]警告: 需要安装openpyxl和pandas库[/yellow]")
            return ""
        
        filepath = self.output_dir / f"{filename}{self.get_extension()}"
        
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. 概览页
        ws_overview = wb.active
        ws_overview.title = "概览"
        self._create_overview_sheet(ws_overview, data, header_font, header_fill, thin_border)
        
        # 2. Commit分析页
        if 'commit_analysis' in data:
            ws_commits = wb.create_sheet("Commit分析")
            self._create_commit_sheet(ws_commits, data['commit_analysis'], 
                                      header_font, header_fill, thin_border)
        
        # 3. 贡献者分析页
        if 'contributor_analysis' in data:
            ws_contrib = wb.create_sheet("贡献者分析")
            self._create_contributor_sheet(ws_contrib, data['contributor_analysis'],
                                          header_font, header_fill, thin_border)
        
        # 4. Issue分析页
        if 'issue_analysis' in data:
            ws_issues = wb.create_sheet("Issue分析")
            self._create_issue_sheet(ws_issues, data['issue_analysis'],
                                    header_font, header_fill, thin_border)
        
        # 5. PR分析页
        if 'pr_analysis' in data:
            ws_prs = wb.create_sheet("PR分析")
            self._create_pr_sheet(ws_prs, data['pr_analysis'],
                                 header_font, header_fill, thin_border)
        
        # 6. 原始数据页
        ws_raw = wb.create_sheet("原始数据")
        self._create_raw_data_sheet(ws_raw, data)
        
        wb.save(filepath)
        console.print(f"[green]Excel报告已生成: {filepath}[/green]")
        return str(filepath)
    
    def _create_overview_sheet(self, ws, data: Dict, header_font, header_fill, border):
        """创建概览页"""
        repo_info = data.get('repo_info', {})
        
        ws['A1'] = "📊 GitHub仓库分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "仓库信息"
        ws['A3'].font = Font(bold=True, size=14)
        
        info_rows = [
            ('仓库名称', repo_info.get('full_name', 'N/A')),
            ('描述', repo_info.get('description', 'N/A')),
            ('主要语言', repo_info.get('language', 'N/A')),
            ('Stars', repo_info.get('stars', 0)),
            ('Forks', repo_info.get('forks', 0)),
            ('Watchers', repo_info.get('watchers', 0)),
            ('Open Issues', repo_info.get('open_issues', 0)),
            ('许可证', repo_info.get('license', 'N/A')),
            ('创建时间', str(repo_info.get('created_at', 'N/A'))[:10]),
            ('最后更新', str(repo_info.get('updated_at', 'N/A'))[:10]),
        ]
        
        for i, (key, value) in enumerate(info_rows, start=5):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # 分析统计摘要
        ws['A17'] = "分析统计摘要"
        ws['A17'].font = Font(bold=True, size=14)
        
        commit_data = data.get('commit_analysis', {})
        contrib_data = data.get('contributor_analysis', {})
        issue_data = data.get('issue_analysis', {})
        pr_data = data.get('pr_analysis', {})
        
        summary_rows = [
            ('总Commits', commit_data.get('total_commits', 0)),
            ('活跃贡献者', contrib_data.get('total_contributors', 0)),
            ('总Issues', issue_data.get('total_issues', 0)),
            ('总PRs', pr_data.get('total_prs', 0)),
        ]
        
        for i, (key, value) in enumerate(summary_rows, start=19):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_commit_sheet(self, ws, commit_data: Dict, header_font, header_fill, border):
        """创建Commit分析页"""
        # 月度分布数据
        monthly = commit_data.get('monthly_distribution', {}).get('distribution', {})
        
        ws['A1'] = "月份"
        ws['B1'] = "Commit数量"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        
        for i, (month, count) in enumerate(monthly.items(), start=2):
            ws[f'A{i}'] = month
            ws[f'B{i}'] = count
        
        # 添加作者统计
        ws['D1'] = "作者"
        ws['E1'] = "Commit数量"
        ws['D1'].font = header_font
        ws['D1'].fill = header_fill
        ws['E1'].font = header_font
        ws['E1'].fill = header_fill
        
        top_authors = commit_data.get('author_stats', {}).get('top_authors', {})
        for i, (author, count) in enumerate(list(top_authors.items())[:20], start=2):
            ws[f'D{i}'] = author
            ws[f'E{i}'] = count
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
    
    def _create_contributor_sheet(self, ws, contrib_data: Dict, header_font, header_fill, border):
        """创建贡献者分析页"""
        headers = ['排名', '用户名', '贡献数', '占比(%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        contributors = contrib_data.get('contributors', [])
        for i, contrib in enumerate(contributors[:50], start=2):
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=contrib.get('login', ''))
            ws.cell(row=i, column=3, value=contrib.get('contributions', 0))
            ws.cell(row=i, column=4, value=round(contrib.get('percentage', 0), 2))
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
    
    def _create_issue_sheet(self, ws, issue_data: Dict, header_font, header_fill, border):
        """创建Issue分析页"""
        ws['A1'] = "Issue统计"
        ws['A1'].font = Font(bold=True, size=14)
        
        stats = [
            ('总Issues', issue_data.get('total_issues', 0)),
            ('开放Issues', issue_data.get('open_issues', 0)),
            ('关闭Issues', issue_data.get('closed_issues', 0)),
            ('平均关闭时间(天)', round(issue_data.get('avg_close_time_days', 0), 2)),
        ]
        
        for i, (key, value) in enumerate(stats, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # 标签统计
        ws['D1'] = "标签统计"
        ws['D1'].font = Font(bold=True, size=14)
        
        ws['D3'] = "标签"
        ws['E3'] = "数量"
        ws['D3'].font = header_font
        ws['D3'].fill = header_fill
        ws['E3'].font = header_font
        ws['E3'].fill = header_fill
        
        labels = issue_data.get('labels', {})
        for i, (label, count) in enumerate(list(labels.items())[:20], start=4):
            ws[f'D{i}'] = label
            ws[f'E{i}'] = count
    
    def _create_pr_sheet(self, ws, pr_data: Dict, header_font, header_fill, border):
        """创建PR分析页"""
        ws['A1'] = "Pull Request统计"
        ws['A1'].font = Font(bold=True, size=14)
        
        stats = [
            ('总PRs', pr_data.get('total_prs', 0)),
            ('开放PRs', pr_data.get('open_prs', 0)),
            ('合并PRs', pr_data.get('merged_prs', 0)),
            ('关闭PRs', pr_data.get('closed_prs', 0)),
            ('合并率(%)', round(pr_data.get('merge_rate', 0), 2)),
            ('平均审查时间(天)', round(pr_data.get('avg_review_time_days', 0), 2)),
        ]
        
        for i, (key, value) in enumerate(stats, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
    
    def _create_raw_data_sheet(self, ws, data: Dict):
        """创建原始数据页"""
        ws['A1'] = "原始JSON数据"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = json.dumps(data, indent=2, default=str, ensure_ascii=False)


class MarkdownExporter(ExportStrategy):
    """Markdown导出器"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def get_extension(self) -> str:
        return ".md"
    
    def export(self, data: Dict, filename: str) -> str:
        """导出为Markdown格式"""
        filepath = self.output_dir / f"{filename}{self.get_extension()}"
        
        content = []
        
        # 标题
        repo_info = data.get('repo_info', {})
        content.append(f"# 📊 {repo_info.get('full_name', 'Unknown')} 分析报告\n")
        content.append(f"> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # 仓库概览
        content.append("## 📋 仓库概览\n")
        content.append(f"| 属性 | 值 |")
        content.append(f"|------|-----|")
        content.append(f"| 仓库名称 | {repo_info.get('full_name', 'N/A')} |")
        content.append(f"| 描述 | {repo_info.get('description', 'N/A')} |")
        content.append(f"| 主要语言 | {repo_info.get('language', 'N/A')} |")
        content.append(f"| ⭐ Stars | {repo_info.get('stars', 0):,} |")
        content.append(f"| 🍴 Forks | {repo_info.get('forks', 0):,} |")
        content.append(f"| 📜 许可证 | {repo_info.get('license', 'N/A')} |")
        content.append("")
        
        # Commit分析
        commit_data = data.get('commit_analysis', {})
        if commit_data:
            content.append("## 📝 Commit分析\n")
            content.append(f"- **总Commits**: {commit_data.get('total_commits', 0):,}")
            content.append(f"- **分析作者数**: {commit_data.get('author_stats', {}).get('total_authors', 0)}")
            content.append("")
            
            # Top作者
            content.append("### 🏆 Top贡献者\n")
            content.append("| 排名 | 作者 | Commits |")
            content.append("|------|------|---------|")
            top_authors = commit_data.get('author_stats', {}).get('top_authors', {})
            for i, (author, count) in enumerate(list(top_authors.items())[:10], start=1):
                content.append(f"| {i} | {author} | {count} |")
            content.append("")
        
        # 贡献者分析
        contrib_data = data.get('contributor_analysis', {})
        if contrib_data:
            content.append("## 👥 贡献者分析\n")
            content.append(f"- **总贡献者**: {contrib_data.get('total_contributors', 0)}")
            content.append("")
        
        # Issue分析
        issue_data = data.get('issue_analysis', {})
        if issue_data:
            content.append("## 🐛 Issue分析\n")
            content.append(f"- **总Issues**: {issue_data.get('total_issues', 0)}")
            content.append(f"- **开放Issues**: {issue_data.get('open_issues', 0)}")
            content.append(f"- **关闭Issues**: {issue_data.get('closed_issues', 0)}")
            content.append(f"- **平均关闭时间**: {issue_data.get('avg_close_time_days', 0):.1f} 天")
            content.append("")
        
        # PR分析
        pr_data = data.get('pr_analysis', {})
        if pr_data:
            content.append("## 🔀 Pull Request分析\n")
            content.append(f"- **总PRs**: {pr_data.get('total_prs', 0)}")
            content.append(f"- **合并PRs**: {pr_data.get('merged_prs', 0)}")
            content.append(f"- **合并率**: {pr_data.get('merge_rate', 0):.1f}%")
            content.append("")
        
        # 写入文件
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content))
        
        console.print(f"[green]Markdown报告已生成: {filepath}[/green]")
        return str(filepath)


class PDFExporter(ExportStrategy):
    """PDF导出器"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def get_extension(self) -> str:
        return ".pdf"
    
    def export(self, data: Dict, filename: str) -> str:
        """导出为PDF格式"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            console.print("[yellow]警告: 需要安装reportlab库来生成PDF[/yellow]")
            # 回退到Markdown
            return MarkdownExporter(str(self.output_dir)).export(data, filename)
        
        filepath = self.output_dir / f"{filename}{self.get_extension()}"
        
        # 创建PDF文档
        doc = SimpleDocTemplate(str(filepath), pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        # 获取样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading1'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10
        )
        
        elements = []
        
        # 标题
        repo_info = data.get('repo_info', {})
        elements.append(Paragraph(f"GitHub仓库分析报告", title_style))
        elements.append(Paragraph(f"仓库: {repo_info.get('full_name', 'Unknown')}", styles['Normal']))
        elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # 仓库概览表格
        elements.append(Paragraph("仓库概览", heading_style))
        overview_data = [
            ['属性', '值'],
            ['仓库名称', repo_info.get('full_name', 'N/A')],
            ['主要语言', repo_info.get('language', 'N/A')],
            ['Stars', str(repo_info.get('stars', 0))],
            ['Forks', str(repo_info.get('forks', 0))],
            ['许可证', repo_info.get('license', 'N/A')],
        ]
        
        overview_table = Table(overview_data, colWidths=[4*cm, 10*cm])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(overview_table)
        elements.append(Spacer(1, 20))
        
        # Commit统计
        commit_data = data.get('commit_analysis', {})
        if commit_data:
            elements.append(Paragraph("Commit分析", heading_style))
            commit_stats = [
                ['指标', '数值'],
                ['总Commits', str(commit_data.get('total_commits', 0))],
                ['作者数', str(commit_data.get('author_stats', {}).get('total_authors', 0))],
            ]
            commit_table = Table(commit_stats, colWidths=[6*cm, 6*cm])
            commit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(commit_table)
            elements.append(Spacer(1, 20))
        
        # Issue统计
        issue_data = data.get('issue_analysis', {})
        if issue_data:
            elements.append(Paragraph("Issue分析", heading_style))
            issue_stats = [
                ['指标', '数值'],
                ['总Issues', str(issue_data.get('total_issues', 0))],
                ['开放Issues', str(issue_data.get('open_issues', 0))],
                ['关闭Issues', str(issue_data.get('closed_issues', 0))],
            ]
            issue_table = Table(issue_stats, colWidths=[6*cm, 6*cm])
            issue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(issue_table)
        
        # 生成PDF
        doc.build(elements)
        console.print(f"[green]PDF报告已生成: {filepath}[/green]")
        return str(filepath)


class ExportManager:
    """导出管理器"""
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        self._exporters = {
            'csv': CSVExporter(output_dir),
            'excel': ExcelExporter(output_dir),
            'xlsx': ExcelExporter(output_dir),
            'markdown': MarkdownExporter(output_dir),
            'md': MarkdownExporter(output_dir),
            'pdf': PDFExporter(output_dir),
        }
    
    def export(self, data: Dict, filename: str, format: str = 'excel') -> str:
        """
        导出分析结果
        
        Args:
            data: 分析结果数据
            filename: 文件名（不含扩展名）
            format: 导出格式 ('csv', 'excel', 'markdown', 'pdf')
            
        Returns:
            导出文件路径
        """
        format_lower = format.lower()
        
        if format_lower not in self._exporters:
            raise ValueError(f"不支持的导出格式: {format}. 支持的格式: {list(self._exporters.keys())}")
        
        exporter = self._exporters[format_lower]
        return exporter.export(data, filename)
    
    def export_all(self, data: Dict, filename: str) -> Dict[str, str]:
        """
        导出所有支持的格式
        
        Args:
            data: 分析结果数据
            filename: 文件名（不含扩展名）
            
        Returns:
            各格式对应的文件路径
        """
        results = {}
        for format_name in ['csv', 'excel', 'markdown', 'pdf']:
            try:
                results[format_name] = self.export(data, filename, format_name)
            except Exception as e:
                console.print(f"[yellow]导出{format_name}失败: {e}[/yellow]")
                results[format_name] = None
        
        return results
    
    def register_exporter(self, name: str, exporter: ExportStrategy):
        """注册自定义导出器"""
        self._exporters[name.lower()] = exporter
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的导出格式列表"""
        return list(set(self._exporters.keys()))


class BatchExporter:
    """批量导出器 - 支持多仓库批量导出"""
    
    def __init__(self, output_dir: str = "output/batch"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.export_manager = ExportManager(str(self.output_dir))
    
    def export_multiple(self, repos_data: List[Dict], format: str = 'excel') -> List[str]:
        """
        批量导出多个仓库的分析结果
        
        Args:
            repos_data: 仓库分析结果列表
            format: 导出格式
            
        Returns:
            导出文件路径列表
        """
        exported_files = []
        
        for data in repos_data:
            repo_name = data.get('repo_info', {}).get('full_name', 'unknown')
            safe_name = repo_name.replace('/', '_')
            
            try:
                filepath = self.export_manager.export(data, safe_name, format)
                exported_files.append(filepath)
                console.print(f"[green]✓ 导出成功: {repo_name}[/green]")
            except Exception as e:
                console.print(f"[red]✗ 导出失败 {repo_name}: {e}[/red]")
        
        return exported_files
    
    def export_comparison_report(self, repos_data: List[Dict], filename: str = "comparison") -> str:
        """
        导出多仓库对比报告
        
        Args:
            repos_data: 仓库分析结果列表
            filename: 文件名
            
        Returns:
            导出文件路径
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            console.print("[yellow]需要安装pandas和openpyxl[/yellow]")
            return ""
        
        filepath = self.output_dir / f"{filename}.xlsx"
        
        # 提取对比数据
        comparison_data = []
        for data in repos_data:
            repo_info = data.get('repo_info', {})
            commit_data = data.get('commit_analysis', {})
            contrib_data = data.get('contributor_analysis', {})
            issue_data = data.get('issue_analysis', {})
            pr_data = data.get('pr_analysis', {})
            
            comparison_data.append({
                '仓库': repo_info.get('full_name', 'N/A'),
                '语言': repo_info.get('language', 'N/A'),
                'Stars': repo_info.get('stars', 0),
                'Forks': repo_info.get('forks', 0),
                'Commits': commit_data.get('total_commits', 0),
                '贡献者': contrib_data.get('total_contributors', 0),
                'Issues': issue_data.get('total_issues', 0),
                'PRs': pr_data.get('total_prs', 0),
                'PR合并率(%)': round(pr_data.get('merge_rate', 0), 2),
            })
        
        # 创建DataFrame并导出
        df = pd.DataFrame(comparison_data)
        df.to_excel(filepath, index=False, sheet_name='仓库对比')
        
        console.print(f"[green]对比报告已生成: {filepath}[/green]")
        return str(filepath)
